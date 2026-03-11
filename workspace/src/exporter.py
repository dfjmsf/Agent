import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from typing import List, Dict, Any, Optional
import os


class RecipeExporter:
    """
    泰拉瑞亚合成表数据导出器类
    
    提供将解析后的合成表数据导出到Excel文件的功能
    """
    
    def __init__(self):
        """初始化导出器"""
        pass
    
    def export_to_excel(self, recipes_data: List[Dict[str, Any]], output_path: str) -> bool:
        """
        将合成表数据导出到Excel文件
        
        Args:
            recipes_data: 包含合成配方数据的字典列表
            output_path: 输出Excel文件的路径
            
        Returns:
            bool: 导出成功返回True，失败返回False
        """
        try:
            # 创建一个新的工作簿
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Recipes'
            
            # 设置表头
            headers = ['Item', 'Materials', 'Station']
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 填充数据
            for row_num, recipe in enumerate(recipes_data, 2):
                # Item列
                item_value = recipe.get('item_name', '')
                if isinstance(item_value, (list, tuple)):
                    item_value = ', '.join(str(x) for x in item_value)
                else:
                    item_value = str(item_value) if item_value is not None else ''
                
                sheet.cell(row=row_num, column=1, value=item_value)
                
                # Materials列
                materials_value = recipe.get('materials', [])
                if isinstance(materials_value, list):
                    materials_str = ', '.join(str(material) for material in materials_value)
                else:
                    materials_str = str(materials_value) if materials_value is not None else ''
                
                sheet.cell(row=row_num, column=2, value=materials_str)
                
                # Station列
                station_value = recipe.get('station', '')
                if isinstance(station_value, (list, tuple)):
                    station_value = ', '.join(str(x) for x in station_value)
                else:
                    station_value = str(station_value) if station_value is not None else ''
                
                sheet.cell(row=row_num, column=3, value=station_value)
            
            # 调整列宽
            sheet.column_dimensions['A'].width = 30  # Item列
            sheet.column_dimensions['B'].width = 50  # Materials列
            sheet.column_dimensions['C'].width = 20  # Station列
            
            # 保存文件
            workbook.save(output_path)
            return True
            
        except Exception as e:
            print(f"导出Excel文件时发生错误: {e}")
            return False


def export_recipes_to_excel(recipes_data: List[Dict[str, Any]], output_path: str) -> bool:
    """
    将合成表数据导出到Excel文件的便捷函数
    
    Args:
        recipes_data: 包含合成配方数据的字典列表
        output_path: 输出Excel文件的路径
        
    Returns:
        bool: 导出成功返回True，失败返回False
    """
    exporter = RecipeExporter()
    return exporter.export_to_excel(recipes_data, output_path)


def export_recipes_to_excel_with_pandas(recipes_data: List[Dict[str, Any]], output_path: str) -> bool:
    """
    使用pandas将合成表数据导出到Excel文件
    
    Args:
        recipes_data: 包含合成配方数据的字典列表
        output_path: 输出Excel文件的路径
        
    Returns:
        bool: 导出成功返回True，失败返回False
    """
    try:
        # 准备数据用于DataFrame
        formatted_data = []
        for recipe in recipes_data:
            item_value = recipe.get('item_name', '')
            if isinstance(item_value, (list, tuple)):
                item_value = ', '.join(str(x) for x in item_value)
            else:
                item_value = str(item_value) if item_value is not None else ''
            
            materials_value = recipe.get('materials', [])
            if isinstance(materials_value, list):
                materials_str = ', '.join(str(material) for material in materials_value)
            else:
                materials_str = str(materials_value) if materials_value is not None else ''
            
            station_value = recipe.get('station', '')
            if isinstance(station_value, (list, tuple)):
                station_value = ', '.join(str(x) for x in station_value)
            else:
                station_value = str(station_value) if station_value is not None else ''
            
            formatted_data.append({
                'Item': item_value,
                'Materials': materials_str,
                'Station': station_value
            })
        
        # 创建DataFrame
        df = pd.DataFrame(formatted_data, columns=['Item', 'Materials', 'Station'])
        
        # 写入Excel文件
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Recipes', index=False)
            
            # 获取工作表对象以调整格式
            worksheet = writer.sheets['Recipes']
            
            # 调整列宽
            worksheet.column_dimensions['A'].width = 30  # Item列
            worksheet.column_dimensions['B'].width = 50  # Materials列
            worksheet.column_dimensions['C'].width = 20  # Station列
        
        return True
        
    except Exception as e:
        print(f"使用pandas导出Excel文件时发生错误: {e}")
        return False


if __name__ == "__main__":
    # 示例用法
    sample_recipes = [
        {
            'item_name': 'Wooden Sword',
            'materials': ['Wood', 'Wood'],
            'station': 'Work Bench'
        },
        {
            'item_name': 'Iron Pickaxe',
            'materials': ['Iron Bar', 'Iron Bar', 'Iron Bar', 'Wood'],
            'station': 'Anvil'
        }
    ]
    
    output_file = "terraria_recipes.xlsx"
    success = export_recipes_to_excel(sample_recipes, output_file)
    
    if success:
        print(f"配方数据已成功导出到 {output_file}")
    else:
        print("配方数据导出失败")